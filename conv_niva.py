import csv
import json
from datetime import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
import requests
import json


filename = "events_2023-03-13.csv"
data = []



# Open the CSV file and read the data into a list
with open(filename, "r") as file:
    reader = csv.reader(file)
    for row in reader:
        data.append(row)

# Extract the required information into a new list
# print only first 5 rows

#function to convert string to dictionary
def convert_to_dict(string):
    #remove the curly braces
    string = string[1:-1]
    #split the string into a list of key-value pairs
    pairs = string.split(",")
    #create a dictionary from the list of key-value pairs
    dictionary = {}
    for pair in pairs:
        key_value = pair.split(":")
        key = key_value[0]
        value = key_value[1]
        dictionary[key] = value
    return dictionary


events_data = []
# taking only 90000 to 100000 rows
# for row in data[106382:125701]:
for row in data[125700:129921]:
    # print(row[4], row[7], row[12], row[13])
    #append to events_data
    # print(row[12])
    events_data.append({"target":row[4], "type":row[7], "event":row[12], "createdOn":row[13]})

# function to fetch queue id from the target id
def get_queue_id(target_id):
    try:
        url = "https://botpress.prod.niva.aws.pncloud.se/api/v1/bots/postnord/mod/clearIt/api/routing-api"
        payload = {"Variables": {"IVA_CALL_ID": target_id}}
        headers = {
            'Content-Type': 'application/json',
            'Cookie': 'AWSALB=nMsiZp9/1gwUvVn8FilDdUKXFR0i+3jcM7c3/zwvcar2XByAxLOdWhHFO1ohZN84ynCmacHNru6jRf2AA2pZ+vWphbAtd/q/1I7vjIIvif00JPhzktXjwpMMv0MH; AWSALBCORS=nMsiZp9/1gwUvVn8FilDdUKXFR0i+3jcM7c3/zwvcar2XByAxLOdWhHFO1ohZN84ynCmacHNru6jRf2AA2pZ+vWphbAtd/q/1I7vjIIvif00JPhzktXjwpMMv0MH'
            }
        response  = requests.request("POST", url, headers=headers, data=json.dumps(payload))
        # print("Queue id for target id", target_id, "is", response.json()["Variables"]["queueGuid"])
        return response.json()["Variables"]["queueGuid"]
    except:
        print("Error: Could not fetch queue id for target id", target_id)
        return "ERROR"




# function to combine all events into one list if they have the same target
def combine_events(events_data):
    if events_data is []:
        return []
    
    # combine events with same target
    combined_events = []
    for event in events_data:
        if event["target"] in [x["target"] for x in combined_events]:
            # find the index of the target in combined_events
            index = [x["target"] for x in combined_events].index(event["target"])
            # append the event to the list of events
            combined_events[index]["events"].append(event["event"])
        else:
            # create a new entry in combined_events
            combined_events.append({"target":event["target"], "events":[event["event"]]})
        
    
    return combined_events

combined_target=combine_events(events_data)

# event = {"type": "text", "channel": "api", "direction": "incoming", "payload": {"type": "text", "text": "privatperson", "phoneNumber": "+46761898100"}, "target": "266987947161210216537822477879871288411", "botId": "postnord", "createdOn": "2023-03-12T06:46:14.452Z", "id": "1700525249523298907", "preview": "privatperson", "flags": {}, "suggestions": [], "nlu": {"entities": [], "language": "sv", "detectedLanguage": "sv", "ambiguous": False, "slots": {}, "intent": {"confidence": 1, "context": "entry", "name": "consumerconfirm"}, "intents": [{"confidence": 1, "context": "entry", "name": "consumerconfirm"}], "errored": False, "includedContexts": [], "ms": 0}, "decision": {"decision": {"reason": "no suggestion matched", "status": "elected"}, "confidence": 1, "payloads": [], "source": "decisionEngine", "sourceDetails": "execute default flow"}}

# function to design the conversation view for each target based on the events
def conversation_view(combined_target):

    if(combined_target is []):
        return []
    
    # create a list of conversation views basedon the createdOn field and direction field
    conversation_views = []
    for target in combined_target:
        # create a list of events for each target
        events = []
        for event in target["events"]:
            
            event_dict = {}
            #convert the event to a dictionary if is decoded as a string
            try:
                event_dict = json.loads(event)

            except json.decoder.JSONDecodeError:
                print("Error: event is not a valid JSON string",event)
            

            # If payload is a dictionary which is not empty and it has a text field called type with a value of text then add the value of the text field to the payload field
            if event_dict and isinstance(event_dict["payload"], dict) and "type" in event_dict["payload"] and event_dict["payload"]["type"] == "text":
                 
                 # convert createdOn to a hour and minute format
                 created_on = datetime.strptime(event_dict["createdOn"], "%Y-%m-%dT%H:%M:%S.%fZ").strftime("%H:%M")

                 events.append({"createdOn":created_on, "direction":event_dict["direction"], "payload":event_dict["payload"]["text"], "intent":event_dict["nlu"]["intent"]["name"]})
           
        
        # sort the event_dict based on the createdOn field
        events = sorted(events, key=lambda k: k["createdOn"])
        
        # create a conversation view for each target
        queue_id = get_queue_id(target["target"])
        conversation_views.append({"target":target["target"], "queue_id": queue_id , "events":events})

    return conversation_views

conversations= conversation_view(combined_target)


# curl --location 'https://botpress.prod.niva.aws.pncloud.se/api/v1/bots/postnord/mod/clearIt/api/routing-api' \
# --header 'Content-Type: application/json' \
# --header 'Cookie: AWSALB=nMsiZp9/1gwUvVn8FilDdUKXFR0i+3jcM7c3/zwvcar2XByAxLOdWhHFO1ohZN84ynCmacHNru6jRf2AA2pZ+vWphbAtd/q/1I7vjIIvif00JPhzktXjwpMMv0MH; AWSALBCORS=nMsiZp9/1gwUvVn8FilDdUKXFR0i+3jcM7c3/zwvcar2XByAxLOdWhHFO1ohZN84ynCmacHNru6jRf2AA2pZ+vWphbAtd/q/1I7vjIIvif00JPhzktXjwpMMv0MH' \
# --data '{
#     "Variables": {
#         "IVA_CALL_ID": "258411551686106911702365268350364894253"
#     }
# }'

# Response
# {
#     "Variables": {
#         "queueGuid": "92d84637-73cf-4113-81a5-c112cf00bfc0",
#         "transcriptions": "NA",
#         "itemId": "21118876328SE"
#     }
# }







# EXPORT 


# Create an Excel workbook and select the active worksheet
wb = openpyxl.Workbook()
ws = wb.active

# Define the column headers and write them to the first row of the worksheet
headers = ["Target ID", "Direction","Intent", "Message", "Time"]
ws.append(headers)

# Define the color and font for the target ID, caller, and bot messages
target_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
caller_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
bot_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
queue_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
font = Font(name="Calibri", size=11)

# Loop through the chat conversations and write them to the worksheet
for conv in conversations:
    target_id = conv["target"]
    events = conv["events"]
    queue_id= conv["queue_id"]

    # Add a row for the target ID
    target_row = ["","Conversation ID", target_id, "", ""]
    queue_row = ["","Queue ID", queue_id, "", ""]
    ws.append(target_row)
    ws.append(queue_row)
    # Fill the row with the target ID and queue ID
    for cell in ws[ws.max_row]:
        cell.font = font
        cell.fill = target_fill
    for cell in ws[ws.max_row-1]:
        cell.font = font
        cell.fill = queue_fill



    # Add a row for each message in the conversation
    for event in events:
        direction = event["direction"]
        message = event["payload"]
        createdOn = event["createdOn"]
        intent = event["intent"]

        # Add formatting to the cell based on the message direction
        if direction == "incoming":
            fill = caller_fill
        else:
            fill = bot_fill

        # Write the direction and message to the row
        row = ["", direction, intent, message, createdOn, ""]
        ws.append(row)

        if ws.max_row > 0:
            for cell in ws[ws.max_row]:
                cell.font = font
                cell.fill = fill

    # Add an empty row to separate conversations
    ws.append([])

# Remove the first column (target ID) from the worksheet
ws.delete_cols(1)

# Save the Excel file
# wb.save("chat_conversations_13_03_2023_15:OO-17:00.xlsx")
wb.save("chat_conversations_05_PM_06_PM.xlsx")