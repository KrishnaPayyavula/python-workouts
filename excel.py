import pandas as pd
from openpyxl import load_workbook

# Load the Excel file using openpyxl engine
book = load_workbook("SampleData.xlsx", read_only=True)

# Select the sheet to read
sheet = book["SalesOrders"]

# Initialize an empty list to store rows with yellow background
yellow_rows = []
white_rows = []

# Loop through each row in the sheet
for row in sheet.iter_rows(min_row=2):  # Skip the header row

    #print the background color of each cell in the row
    print([cell.fill.start_color.index for cell in row])

    # Check if any cell in the row has yellow background color
    if any(cell.fill.start_color.index == "FFFFFF00" for cell in row):
       
        # If yes, append the row to the yellow_rows list
        yellow_rows.append([cell.value for cell in row])
    
    # Check if any cell in the row has white background color (Default color)
    if any(cell.fill.start_color.index == "00000000" for cell in row):
       
        # If yes, append the row to the yellow_rows list
        white_rows.append([cell.value for cell in row])

# Create a Pandas DataFrame from the yellow_rows list
df = pd.DataFrame(yellow_rows, columns=["OrderDate", "Region", "Rep", "Item", "Units", "UnitCost", "Total"])
dff = pd.DataFrame(white_rows, columns=["OrderDate", "Region", "Rep", "Item", "Units", "UnitCost", "Total"])

# Print the DataFrame
print(df)
print("Second Dataframe ********")
print(dff)
